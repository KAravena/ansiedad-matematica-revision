# -*- coding: utf-8 -*-
"""
Pipeline reproducible PRISMA para revisión sistemática exploratoria:
"Ansiedad matemática en estudiantes escolares: definiciones, mediciones y enfoques de análisis..."

Este script organiza, diagnostica, unifica, normaliza, deduplica y prepara matrices
para cribado humano. NO decide automáticamente el corpus final.

Modos:
    python prisma_pipeline_ansiedad_matematica.py --mode prepare
    python prisma_pipeline_ansiedad_matematica.py --mode update_screening
    python prisma_pipeline_ansiedad_matematica.py --mode finalize

Rutas por defecto pensadas para Windows. Puedes sobrescribir con --root, --eric, --scopus, --wos.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import shutil
import sys
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from unidecode import unidecode
except Exception:  # pragma: no cover
    def unidecode(x: str) -> str:
        return x

try:
    from rapidfuzz import fuzz
    HAS_RAPIDFUZZ = True
except Exception:  # pragma: no cover
    fuzz = None
    HAS_RAPIDFUZZ = False

# -----------------------------------------------------------------------------
# Configuración principal
# -----------------------------------------------------------------------------
DEFAULT_ROOT = Path(r"D:\Psicologia\Seminario de investigación\presentación")
DEFAULT_FILES = {
    "ERIC": DEFAULT_ROOT / "ERIC" / "ERIC2026-06-11_18.02.50.csv",
    "Scopus": DEFAULT_ROOT / "SCOPUS" / "scopus_export_Jun 11-2026_0530af68-403f-432f-892c-ad3cada87cb2.csv",
    "WoS": DEFAULT_ROOT / "wos" / "savedrecs.csv",
}
EXPECTED_COUNTS = {"ERIC": 118, "Scopus": 223, "WoS": 240}

FOLDER_NAMES = [
    "01_raw_exports",
    "02_diagnostics",
    "03_unified_database",
    "04_deduplication",
    "05_title_abstract_screening",
    "06_full_text_screening",
    "07_prisma_outputs",
    "08_final_corpus",
    "09_logs",
    "10_scripts_backup",
]

STANDARD_COLUMNS = [
    "id_unico", "base_origen", "id_original", "titulo", "abstract", "autores",
    "año", "revista_fuente", "doi", "idioma", "tipo_documento", "palabras_clave",
    "url", "texto_screening", "titulo_normalizado", "doi_limpio", "tiene_doi",
    "tiene_abstract", "longitud_abstract", "anio_valido_2021_2026",
    "requiere_verificacion_anio", "fila_anomala_si_no", "notas_anomalia",
]

EXCLUSION_CODES = {
    "R1": "Población no escolar",
    "R2": "Educación superior",
    "R3": "Formación docente",
    "R4": "No mide ansiedad matemática",
    "R5": "No se relaciona con aprendizaje/rendimiento/experiencia educativa",
    "R6": "No empírico",
    "R7": "Revisión/metaanálisis",
    "R8": "Documento no artículo",
    "R9": "Sin abstract suficiente",
    "R10": "Sin texto completo",
    "R11": "Duplicado",
}

# -----------------------------------------------------------------------------
# Utilidades generales
# -----------------------------------------------------------------------------
def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_dirs(root: Path) -> Dict[str, Path]:
    dirs = {name: root / name for name in FOLDER_NAMES}
    for path in dirs.values():
        path.mkdir(parents=True, exist_ok=True)
    (dirs["06_full_text_screening"] / "pdfs").mkdir(parents=True, exist_ok=True)
    return dirs


def setup_logging(log_dir: Path, mode: str) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / "pipeline_log.txt"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(log_path, encoding="utf-8", mode="a"), logging.StreamHandler(sys.stdout)],
    )
    logging.info("=" * 90)
    logging.info("Inicio pipeline | modo=%s", mode)


def safe_copy(src: Path, dst_dir: Path) -> Path:
    if not src.exists():
        logging.warning("No existe archivo original: %s", src)
        return dst_dir / src.name
    dst = dst_dir / src.name
    if dst.exists():
        stem, suffix = dst.stem, dst.suffix
        dst = dst_dir / f"{stem}_copy_{timestamp()}{suffix}"
    shutil.copy2(src, dst)
    logging.info("Archivo respaldado: %s -> %s", src, dst)
    return dst


def versioned_path(path: Path) -> Path:
    """Evita sobrescribir outputs agregando timestamp si el archivo ya existe."""
    if not path.exists():
        return path
    return path.with_name(f"{path.stem}_{timestamp()}{path.suffix}")


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    text = unidecode(text)
    text = text.lower()
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[^a-z0-9áéíóúñü\s:/._-]", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_title(value) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_doi(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    if not text or text in {"nan", "none", "na"}:
        return ""
    text = text.replace("https://doi.org/", "")
    text = text.replace("http://dx.doi.org/", "")
    text = text.replace("http://doi.org/", "")
    text = re.sub(r"^doi:\s*", "", text)
    text = re.sub(r"\s+", "", text)
    # Si viene mezclado con texto, intentar capturar un DOI real.
    match = re.search(r'10\.\d{4,9}/[^\s;,"]+', text)
    return match.group(0).rstrip(".") if match else text


def coerce_year(value) -> Optional[int]:
    if pd.isna(value):
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "na"}:
        return None
    m = re.search(r"(19|20)\d{2}", s)
    if m:
        return int(m.group(0))
    try:
        return int(float(s))
    except Exception:
        return None


def join_fields(*values) -> str:
    parts = []
    for v in values:
        if v is None or pd.isna(v):
            continue
        s = str(v).strip()
        if s and s.lower() not in {"nan", "none"}:
            parts.append(s)
    return "; ".join(parts)


def first_existing(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    lower_map = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower().strip() in lower_map:
            return lower_map[cand.lower().strip()]
    return None


def get_series(df: pd.DataFrame, candidates: Iterable[str], default="") -> pd.Series:
    col = first_existing(df, candidates)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col].fillna(default)


def combine_series(df: pd.DataFrame, candidates: Iterable[str]) -> pd.Series:
    found = [c for c in candidates if c in df.columns]
    if not found:
        # case insensitive
        lower_map = {c.lower().strip(): c for c in df.columns}
        found = [lower_map[c.lower().strip()] for c in candidates if c.lower().strip() in lower_map]
    if not found:
        return pd.Series([""] * len(df), index=df.index)
    return df[found].fillna("").astype(str).agg("; ".join, axis=1).str.replace(r"(;\s*)+", "; ", regex=True).str.strip("; ")


def text_contains(text: str, terms: Iterable[str]) -> List[str]:
    t = normalize_text(text)
    hits = []
    for term in terms:
        # buscar término normalizado como frase; para términos muy cortos usar frontera de palabra
        term_norm = normalize_text(term)
        if not term_norm:
            continue
        if re.search(r"\b" + re.escape(term_norm) + r"\b", t):
            hits.append(term)
    return hits


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if HAS_RAPIDFUZZ:
        return float(fuzz.ratio(a, b))
    return SequenceMatcher(None, a, b).ratio() * 100.0

# -----------------------------------------------------------------------------
# Lectura y diagnóstico de CSV
# -----------------------------------------------------------------------------
def read_csv_robust(path: Path) -> Tuple[pd.DataFrame, str, str, List[str]]:
    """Lee un CSV probando codificaciones y separadores. Devuelve df, encoding, sep, warnings."""
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
    seps = [",", ";", "\t"]
    warnings: List[str] = []

    best = None
    best_meta = None

    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(
                    path,
                    encoding=enc,
                    sep=sep,
                    engine="python",
                    dtype=str,
                    keep_default_na=False,
                    on_bad_lines="warn",
                    quoting=csv.QUOTE_MINIMAL,
                )
                # Preferir lectura con más columnas y filas razonables.
                score = len(df) * max(1, len(df.columns))
                if best is None or score > best_meta[0]:
                    best = df
                    best_meta = (score, enc, sep)
            except Exception as e:
                warnings.append(f"Fallo lectura {path.name} encoding={enc} sep={repr(sep)}: {e}")
    if best is None:
        raise RuntimeError(f"No se pudo leer el archivo: {path}")

    _, enc, sep = best_meta
    # Quitar columnas completamente vacías o unnamed redundantes solo si no contienen datos.
    best.columns = [str(c).strip() for c in best.columns]
    # columnas duplicadas: pandas agrega .1; conservarlas pero registrar.
    logging.info("CSV leído: %s | encoding=%s | sep=%s | shape=%s", path.name, enc, repr(sep), best.shape)
    return best, enc, sep, warnings


def detect_base_from_name(path: Path) -> str:
    name = path.name.lower()
    if "eric" in name:
        return "ERIC"
    if "scopus" in name:
        return "Scopus"
    if "savedrecs" in name or "wos" in name:
        return "WoS"
    return "Desconocida"


def diagnostic_for_df(df: pd.DataFrame, base: str, path: Path, encoding: str, sep: str) -> Dict:
    if base == "ERIC":
        title_col = first_existing(df, ["Title"])
        abs_col = first_existing(df, ["Abstract Note", "Abstract"])
        doi_col = first_existing(df, ["DOI"])
        year_col = first_existing(df, ["Publication Year", "Year"])
        type_col = first_existing(df, ["Item Type", "Document Type", "Publication Type"])
    elif base == "Scopus":
        title_col = first_existing(df, ["Title"])
        abs_col = first_existing(df, ["Abstract"])
        doi_col = first_existing(df, ["DOI"])
        year_col = first_existing(df, ["Year"])
        type_col = first_existing(df, ["Document Type"])
    else:  # WoS
        title_col = first_existing(df, ["Article Title", "Title"])
        abs_col = first_existing(df, ["Abstract"])
        doi_col = first_existing(df, ["DOI"])
        year_col = first_existing(df, ["Publication Year", "Year"])
        type_col = first_existing(df, ["Document Type", "Publication Type"])

    title_s = df[title_col] if title_col else pd.Series([""] * len(df))
    abs_s = df[abs_col] if abs_col else pd.Series([""] * len(df))
    doi_s = df[doi_col] if doi_col else pd.Series([""] * len(df))
    year_s = df[year_col] if year_col else pd.Series([""] * len(df))
    type_s = df[type_col] if type_col else pd.Series([""] * len(df))

    years_parsed = year_s.apply(coerce_year)
    invalid_year_rows = int((years_parsed.isna() | ~years_parsed.between(2021, 2026)).sum())

    expected_doc_types = {"article", "journalarticle", "journal article", "j", "proceedings paper", "review", "book chapter"}
    type_norm = type_s.apply(normalize_text)
    type_anomaly = type_norm.apply(lambda x: bool(x) and len(x) > 80)

    diag = {
        "base": base,
        "ruta": str(path),
        "archivo": path.name,
        "encoding_usado": encoding,
        "separador_usado": repr(sep),
        "filas_leidas": len(df),
        "columnas_leidas": len(df.columns),
        "columnas": "; ".join(map(str, df.columns)),
        "primeros_3_titulos": " | ".join(title_s.fillna("").astype(str).head(3).tolist()),
        "titulos_vacios": int(title_s.fillna("").astype(str).str.strip().eq("").sum()),
        "abstracts_vacios": int(abs_s.fillna("").astype(str).str.strip().eq("").sum()),
        "doi_vacios": int(doi_s.fillna("").astype(str).str.strip().eq("").sum()),
        "anios_detectados": "; ".join(sorted(set(year_s.fillna("").astype(str).head(30).tolist()))),
        "tipos_documentales_detectados": "; ".join(sorted(set(type_s.fillna("").astype(str).head(30).tolist()))),
        "filas_anio_invalido_o_fuera_rango": invalid_year_rows,
        "filas_tipo_documental_anomalo": int(type_anomaly.sum()),
        "filas_posible_desalineacion": int(type_anomaly.sum()),
    }
    return diag

# -----------------------------------------------------------------------------
# Normalización por base
# -----------------------------------------------------------------------------
def normalize_base(df: pd.DataFrame, base: str) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    out["base_origen"] = base

    if base == "ERIC":
        out["id_original"] = get_series(df, ["Key"])
        out["titulo"] = get_series(df, ["Title"])
        out["abstract"] = get_series(df, ["Abstract Note", "Abstract"])
        out["autores"] = get_series(df, ["Author", "Authors"])
        out["año"] = get_series(df, ["Publication Year", "Year"])
        out["revista_fuente"] = get_series(df, ["Publication Title", "Journal"])
        out["doi"] = get_series(df, ["DOI"])
        out["idioma"] = get_series(df, ["Language"])
        out["tipo_documento"] = get_series(df, ["Item Type", "Document Type"])
        out["palabras_clave"] = combine_series(df, ["Manual Tags", "Automatic Tags", "Extra", "Tags"])
        out["url"] = get_series(df, ["Url", "URL", "Link"])
    elif base == "Scopus":
        out["id_original"] = get_series(df, ["EID"])
        out["titulo"] = get_series(df, ["Title"])
        out["abstract"] = get_series(df, ["Abstract"])
        out["autores"] = get_series(df, ["Authors"])
        out["año"] = get_series(df, ["Year"])
        out["revista_fuente"] = get_series(df, ["Source title"])
        out["doi"] = get_series(df, ["DOI"])
        out["idioma"] = get_series(df, ["Language"])
        out["tipo_documento"] = get_series(df, ["Document Type"])
        out["palabras_clave"] = combine_series(df, ["Author Keywords", "Index Keywords"])
        out["url"] = get_series(df, ["Link", "URL"])
    else:  # WoS
        out["id_original"] = get_series(df, ["UT (Unique ID)", "UT"])
        out["titulo"] = get_series(df, ["Article Title", "Title"])
        out["abstract"] = get_series(df, ["Abstract"])
        out["autores"] = get_series(df, ["Authors"])
        out["año"] = get_series(df, ["Publication Year", "Year"])
        out["revista_fuente"] = get_series(df, ["Source Title", "Source Title - Arabic", "Journal"])
        out["doi"] = get_series(df, ["DOI"])
        out["idioma"] = get_series(df, ["Language"])
        out["tipo_documento"] = get_series(df, ["Document Type", "Publication Type"])
        out["palabras_clave"] = combine_series(df, ["Author Keywords", "Keywords Plus", "Keywords"])
        out["url"] = get_series(df, ["UT (Unique ID)", "UT", "URL"])

    # Incorporar raw extra en texto si aparece en df.
    raw_extra = combine_series(df, ["raw_extra_fields"])
    out["texto_screening"] = (
        out["titulo"].fillna("").astype(str) + " " +
        out["abstract"].fillna("").astype(str) + " " +
        out["palabras_clave"].fillna("").astype(str) + " " + raw_extra.fillna("").astype(str)
    ).str.replace(r"\s+", " ", regex=True).str.strip()

    out["titulo_normalizado"] = out["titulo"].apply(normalize_title)
    out["doi_limpio"] = out["doi"].apply(clean_doi)
    out["tiene_doi"] = np.where(out["doi_limpio"].str.len() > 0, "sí", "no")
    out["tiene_abstract"] = np.where(out["abstract"].fillna("").astype(str).str.strip().str.len() > 0, "sí", "no")
    out["longitud_abstract"] = out["abstract"].fillna("").astype(str).str.len()
    year_num = out["año"].apply(coerce_year)
    out["año"] = year_num
    out["anio_valido_2021_2026"] = np.where(year_num.between(2021, 2026), "sí", "no")
    out["requiere_verificacion_anio"] = np.where(year_num.between(2021, 2026), "no", "sí")

    # Anomalías: tipo documental extremadamente largo o año raro o título vacío.
    type_len = out["tipo_documento"].fillna("").astype(str).str.len()
    title_empty = out["titulo"].fillna("").astype(str).str.strip().eq("")
    anomalous = (type_len > 80) | title_empty | (out["requiere_verificacion_anio"] == "sí")
    out["fila_anomala_si_no"] = np.where(anomalous, "sí", "no")
    notes = []
    for _, row in out.iterrows():
        parts = []
        if len(str(row.get("tipo_documento", ""))) > 80:
            parts.append("tipo_documento anómalo o desalineado")
        if not str(row.get("titulo", "")).strip():
            parts.append("título vacío")
        if row.get("requiere_verificacion_anio") == "sí":
            parts.append("año inválido/fuera de rango")
        notes.append("; ".join(parts))
    out["notas_anomalia"] = notes

    out.insert(0, "id_unico", [f"{base.upper()}_{i+1:04d}" for i in range(len(out))])
    return out[STANDARD_COLUMNS]

# -----------------------------------------------------------------------------
# Exportar Excel con formato básico
# -----------------------------------------------------------------------------
def autosize_and_style(path: Path, freeze: bool = True) -> None:
    try:
        wb = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        thin = Side(border_style="thin", color="D9D9D9")
        for ws in wb.worksheets:
            if freeze:
                ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(bottom=thin)
            for col in ws.columns:
                max_len = 8
                letter = col[0].column_letter
                for cell in col[:200]:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        max_len = max(max_len, min(len(val), 60))
                    except Exception:
                        pass
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 500)):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(path)
    except Exception as e:
        logging.warning("No se pudo aplicar formato a %s: %s", path, e)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], format_file: bool = True) -> Path:
    out = versioned_path(path)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)
    if format_file:
        autosize_and_style(out)
    logging.info("Excel exportado: %s", out)
    return out

# -----------------------------------------------------------------------------
# Deduplicación
# -----------------------------------------------------------------------------
def choose_keeper(group: pd.DataFrame) -> int:
    """Devuelve index del registro a conservar dentro de un grupo duplicado."""
    priority_base = {"WoS": 3, "Scopus": 2, "ERIC": 1}
    temp = group.copy()
    temp["p_doi"] = temp["doi_limpio"].fillna("").astype(str).str.len().gt(0).astype(int)
    temp["p_abs"] = pd.to_numeric(temp["longitud_abstract"], errors="coerce").fillna(0)
    temp["p_base"] = temp["base_origen"].map(priority_base).fillna(0)
    temp = temp.sort_values(["p_doi", "p_abs", "p_base"], ascending=[False, False, False])
    return temp.index[0]


def mark_duplicates(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, int]]:
    data = df.copy()
    data["duplicado_si_no"] = "no"
    data["tipo_duplicado"] = "no"
    data["grupo_duplicado"] = ""
    data["registro_conservado_si_no"] = "sí"
    data["base_origen_combinada"] = data["base_origen"].astype(str)
    data["ids_originales_combinados"] = data["id_original"].fillna("").astype(str)
    data["notas_duplicado"] = ""

    group_id = 0
    duplicate_indices = set()
    keeper_indices = set(data.index.tolist())
    groups_records = []

    def process_groups(grouped, dup_type: str):
        nonlocal group_id, data, duplicate_indices, keeper_indices, groups_records
        for key, group in grouped:
            if not key or len(group) <= 1:
                continue
            # No reprocesar grupos compuestos solo de registros ya duplicados por capa anterior
            group_active = group.loc[~group.index.isin(duplicate_indices)]
            if len(group_active) <= 1:
                continue
            group_id += 1
            gid = f"G{group_id:04d}_{dup_type}"
            keeper = choose_keeper(group_active)
            origins = "; ".join(sorted(set(group_active["base_origen"].dropna().astype(str))))
            ids = "; ".join(sorted(set(group_active["id_original"].fillna("").astype(str))))
            for idx, row in group_active.iterrows():
                data.at[idx, "duplicado_si_no"] = "sí" if idx != keeper else "no"
                data.at[idx, "tipo_duplicado"] = dup_type
                data.at[idx, "grupo_duplicado"] = gid
                data.at[idx, "registro_conservado_si_no"] = "sí" if idx == keeper else "no"
                data.at[idx, "base_origen_combinada"] = origins
                data.at[idx, "ids_originales_combinados"] = ids
                data.at[idx, "notas_duplicado"] = f"Grupo duplicado por {dup_type}; conservar={data.at[keeper, 'id_unico']}"
                groups_records.append({
                    "grupo_duplicado": gid,
                    "tipo_duplicado": dup_type,
                    "clave": key,
                    "id_unico": row["id_unico"],
                    "registro_conservado_si_no": "sí" if idx == keeper else "no",
                    "base_origen": row["base_origen"],
                    "titulo": row["titulo"],
                    "doi_limpio": row["doi_limpio"],
                })
            for idx in group_active.index:
                if idx != keeper:
                    duplicate_indices.add(idx)
                    keeper_indices.discard(idx)

    # Capa 1 DOI exacto
    by_doi = data[data["doi_limpio"].fillna("").astype(str).str.len() > 0].groupby("doi_limpio", dropna=False)
    process_groups(by_doi, "doi_exacto")

    # Capa 2 título exacto solo entre no duplicados; principalmente sin DOI
    active = data.loc[~data.index.isin(duplicate_indices)].copy()
    by_title = active[active["titulo_normalizado"].fillna("").astype(str).str.len() > 0].groupby("titulo_normalizado", dropna=False)
    process_groups(by_title, "titulo_exacto")

    dedup = data[data["registro_conservado_si_no"] == "sí"].copy().reset_index(drop=True)

    # Capa 3 duplicados probables: marcar, no eliminar.
    probable_records = []
    titles = dedup["titulo_normalizado"].fillna("").astype(str).tolist()
    ids = dedup["id_unico"].tolist()
    for i in range(len(dedup)):
        ti = titles[i]
        if len(ti) < 20:
            continue
        for j in range(i + 1, len(dedup)):
            tj = titles[j]
            if len(tj) < 20:
                continue
            # bloquear por primeras letras para reducir comparaciones
            if ti[:8] != tj[:8] and len(set(ti.split()[:4]).intersection(set(tj.split()[:4]))) < 2:
                continue
            score = similarity(ti, tj)
            if score >= 95:
                probable_records.append({
                    "id_unico_1": ids[i],
                    "id_unico_2": ids[j],
                    "similitud_titulo": round(score, 2),
                    "titulo_1": dedup.loc[i, "titulo"],
                    "titulo_2": dedup.loc[j, "titulo"],
                    "doi_1": dedup.loc[i, "doi_limpio"],
                    "doi_2": dedup.loc[j, "doi_limpio"],
                    "decision_manual": "pendiente",
                })
                # marcar ambos como probable en dedup y data
                for idx in [i, j]:
                    current = str(dedup.at[idx, "notas_duplicado"] or "")
                    dedup.at[idx, "tipo_duplicado"] = "titulo_probable"
                    dedup.at[idx, "notas_duplicado"] = (current + "; duplicado probable por similitud").strip("; ")

    groups_df = pd.DataFrame(groups_records)
    probable_df = pd.DataFrame(probable_records)
    summary = {
        "total_antes_deduplicar": len(df),
        "duplicados_eliminados": int((data["registro_conservado_si_no"] == "no").sum()),
        "duplicados_por_doi": int(((data["registro_conservado_si_no"] == "no") & (data["tipo_duplicado"] == "doi_exacto")).sum()),
        "duplicados_por_titulo_exacto": int(((data["registro_conservado_si_no"] == "no") & (data["tipo_duplicado"] == "titulo_exacto")).sum()),
        "duplicados_probables_revision_manual": len(probable_df),
        "registros_unicos_finales": len(dedup),
    }
    return data.reset_index(drop=True), dedup, probable_df, summary | {"groups_df": groups_df}

# -----------------------------------------------------------------------------
# Sugerencias de cribado
# -----------------------------------------------------------------------------
ANXIETY_TERMS = ["math anxiety", "mathematics anxiety", "mathematical anxiety", "ansiedad matematica", "ansiedade matematica"]
SCHOOL_TERMS = ["school", "students", "student", "child", "children", "adolescent", "primary", "secondary", "elementary", "middle school", "high school", "grade", "classroom"]
DIMENSION_TERMS = ["achievement", "performance", "learning", "engagement", "strategy", "strategies", "gender", "socioeconomic", "school context", "classroom", "attainment", "mathematical literacy", "self-efficacy", "attitude", "attitudes"]
REVIEW_TERMS = ["systematic review", "meta-analysis", "meta analysis", "literature review", "scoping review", "narrative review", "review article"]
HIGHER_ED_TERMS = ["university students", "undergraduate", "college students", "higher education"]
TEACHER_ED_TERMS = ["preservice teacher", "pre-service teacher", "teacher education", "teacher candidates", "teachers only"]
NON_EMPIRICAL_TERMS = ["editorial", "commentary", "book chapter", "dissertation", "thesis", "report"]


def auto_suggest(row: pd.Series) -> Tuple[str, str, str, str]:
    text = str(row.get("texto_screening", ""))
    abstract_len = int(row.get("longitud_abstract", 0) or 0)
    year_ok = row.get("anio_valido_2021_2026") == "sí"

    hits_review = text_contains(text, REVIEW_TERMS)
    hits_higher = text_contains(text, HIGHER_ED_TERMS)
    hits_teacher = text_contains(text, TEACHER_ED_TERMS)
    hits_nonemp = text_contains(text, NON_EMPIRICAL_TERMS)

    if hits_review:
        return "excluir_posible", "R7 Revisión/metaanálisis", "; ".join(hits_review), "verificar decisión humana"
    if hits_higher:
        return "excluir_posible", "R2 Educación superior", "; ".join(hits_higher), "verificar decisión humana"
    if hits_teacher:
        return "excluir_posible", "R3 Formación docente", "; ".join(hits_teacher), "verificar decisión humana"
    if hits_nonemp:
        return "excluir_posible", "R6/R8 No empírico o documento no artículo", "; ".join(hits_nonemp), "verificar decisión humana"

    hits_anx = text_contains(text, ANXIETY_TERMS)
    hits_school = text_contains(text, SCHOOL_TERMS)
    hits_dim = text_contains(text, DIMENSION_TERMS)

    criteria_hits = []
    if hits_anx:
        criteria_hits.append("ansiedad=" + ", ".join(hits_anx[:5]))
    if hits_school:
        criteria_hits.append("poblacion=" + ", ".join(hits_school[:5]))
    if hits_dim:
        criteria_hits.append("dimension=" + ", ".join(hits_dim[:5]))

    if not year_ok:
        return "dudoso", "Año inválido o fuera de rango", "; ".join(criteria_hits), "revisar año"
    if abstract_len < 300:
        return "dudoso", "R9 Sin abstract suficiente", "; ".join(criteria_hits), "ante duda conservar"
    if hits_anx and hits_school and hits_dim:
        return "incluir_posible", "cumple bloques mínimos", "; ".join(criteria_hits), "revisar decisión humana"
    return "dudoso", "no cumple claramente inclusión/exclusión", "; ".join(criteria_hits), "ante duda conservar"


def add_screening_suggestions(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    suggestions = out.apply(auto_suggest, axis=1, result_type="expand")
    suggestions.columns = ["sugerencia_auto_titulo_resumen", "razon_sugerencia_auto", "criterios_detectados", "advertencia_revision_manual"]
    out = pd.concat([out, suggestions], axis=1)
    out["decision_titulo_resumen"] = ""
    out["razon_exclusion_titulo_resumen"] = ""
    out["notas_screening"] = ""
    return out

# -----------------------------------------------------------------------------
# Matrices e instrucciones
# -----------------------------------------------------------------------------
def screening_instructions_df() -> pd.DataFrame:
    rows = [
        ("Objetivo", "Esta matriz sirve para cribado humano por título y resumen. La sugerencia automática no es decisión final."),
        ("Incluir", "Usar si el registro es artículo empírico, 2021-2026, población escolar, mide ansiedad matemática y se vincula con una dimensión educativa."),
        ("Excluir", "Usar solo si incumple claramente los criterios. Registrar razón estandarizada."),
        ("Dudoso", "Usar si falta información, el abstract es ambiguo o hay dudas. Ante duda, conservar para texto completo."),
        ("R1", "Población no escolar"),
        ("R2", "Educación superior"),
        ("R3", "Formación docente"),
        ("R4", "No mide ansiedad matemática"),
        ("R5", "No se relaciona con aprendizaje/rendimiento/experiencia educativa"),
        ("R6", "No empírico"),
        ("R7", "Revisión/metaanálisis"),
        ("R8", "Documento no artículo"),
        ("R9", "Sin abstract suficiente"),
        ("R11", "Duplicado"),
    ]
    return pd.DataFrame(rows, columns=["campo", "descripcion"])


def add_validations_to_screening(path: Path) -> None:
    try:
        wb = load_workbook(path)
        ws = wb["matriz_screening"] if "matriz_screening" in wb.sheetnames else wb.worksheets[0]
        headers = [c.value for c in ws[1]]
        if "decision_titulo_resumen" in headers:
            col = headers.index("decision_titulo_resumen") + 1
            dv = DataValidation(type="list", formula1='"incluir,excluir,dudoso"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=max(ws.max_row,2), column=col).coordinate}")
        if "razon_exclusion_titulo_resumen" in headers:
            col = headers.index("razon_exclusion_titulo_resumen") + 1
            formula = '"R1 Población no escolar,R2 Educación superior,R3 Formación docente,R4 No mide ansiedad matemática,R5 No se relaciona,R6 No empírico,R7 Revisión/metaanálisis,R8 Documento no artículo,R9 Sin abstract suficiente,R11 Duplicado"'
            dv2 = DataValidation(type="list", formula1=formula[:250], allow_blank=True)
            ws.add_data_validation(dv2)
            dv2.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=max(ws.max_row,2), column=col).coordinate}")
        wb.save(path)
    except Exception as e:
        logging.warning("No se pudieron agregar validaciones a %s: %s", path, e)


def create_full_text_template(dedup: pd.DataFrame, screening_path: Path, out_path: Path) -> Path:
    if screening_path.exists():
        try:
            screen = pd.read_excel(screening_path, sheet_name="matriz_screening")
            if "decision_titulo_resumen" in screen.columns and screen["decision_titulo_resumen"].fillna("").astype(str).str.len().gt(0).any():
                keep = screen["decision_titulo_resumen"].str.lower().isin(["incluir", "dudoso"])
                base = screen.loc[keep].copy()
            else:
                base = dedup.copy()
        except Exception:
            base = dedup.copy()
    else:
        base = dedup.copy()
    cols = ["id_unico", "titulo", "autores", "año", "revista_fuente", "doi", "url", "base_origen_combinada", "decision_titulo_resumen", "razon_exclusion_titulo_resumen"]
    for c in cols:
        if c not in base.columns:
            base[c] = ""
    out = base[cols].copy()
    out["texto_completo_buscado_si_no"] = ""
    out["texto_completo_recuperado_si_no"] = ""
    out["decision_texto_completo"] = ""
    out["razon_exclusion_texto_completo"] = ""
    out["decision_final"] = ""
    out["notas_texto_completo"] = ""
    notes = pd.DataFrame({"nota": ["Si aún no completaste decision_titulo_resumen, filtra esta matriz después del cribado humano. Casos 'incluir' y 'dudoso' pasan a texto completo."]})
    return write_excel(out_path, {"matriz_texto_completo": out, "instrucciones": notes})


def create_extraction_template(out_path: Path, corpus: Optional[pd.DataFrame] = None) -> Path:
    cols = [
        "id_unico", "referencia_apa", "autores", "año", "titulo", "pais_contexto", "nivel_educativo", "muestra",
        "diseño_metodologico", "tipo_datos", "instrumento_ansiedad_matematica", "como_define_ansiedad_matematica",
        "variable_resultado_o_dimension_asociada", "instrumento_rendimiento_si_aplica", "tratamiento_genero",
        "tratamiento_nse", "tratamiento_contexto_escolar", "estrategia_analitica", "hallazgos_principales",
        "categoria_analisis_1", "categoria_analisis_2", "aporte_para_revision", "limitaciones_reportadas", "notas",
    ]
    if corpus is None or corpus.empty:
        df = pd.DataFrame(columns=cols)
    else:
        df = pd.DataFrame(columns=cols)
        for c in ["id_unico", "autores", "año", "titulo"]:
            if c in corpus.columns:
                df[c] = corpus[c]
    return write_excel(out_path, {"matriz_extraccion": df})

# -----------------------------------------------------------------------------
# PRISMA
# -----------------------------------------------------------------------------
def build_prisma_table(counts: Dict) -> pd.DataFrame:
    rows = [
        ("Identificación", "ERIC esperado", counts.get("expected_ERIC", EXPECTED_COUNTS["ERIC"])),
        ("Identificación", "Scopus esperado", counts.get("expected_Scopus", EXPECTED_COUNTS["Scopus"])),
        ("Identificación", "WoS esperado", counts.get("expected_WoS", EXPECTED_COUNTS["WoS"])),
        ("Identificación", "Total esperado", counts.get("expected_total", sum(EXPECTED_COUNTS.values()))),
        ("Identificación", "ERIC leído", counts.get("read_ERIC", "[pendiente]")),
        ("Identificación", "Scopus leído", counts.get("read_Scopus", "[pendiente]")),
        ("Identificación", "WoS leído", counts.get("read_WoS", "[pendiente]")),
        ("Identificación", "Total leído/exportado", counts.get("read_total", "[pendiente]")),
        ("Deduplicación", "Duplicados eliminados", counts.get("duplicates_removed", "[pendiente]")),
        ("Deduplicación", "Registros después de eliminar duplicados", counts.get("unique_after_dedup", "[pendiente]")),
        ("Cribado", "Registros cribados por título y resumen", counts.get("screened_title_abstract", "[pendiente]")),
        ("Cribado", "Registros excluidos por título y resumen", counts.get("excluded_title_abstract", "[pendiente]")),
        ("Cribado", "Registros enviados a texto completo", counts.get("sent_full_text", "[pendiente]")),
        ("Recuperación", "Textos completos no recuperados", counts.get("full_text_not_retrieved", "[pendiente]")),
        ("Elegibilidad", "Textos completos evaluados", counts.get("full_text_assessed", "[pendiente]")),
        ("Elegibilidad", "Textos completos excluidos con razones", counts.get("full_text_excluded", "[pendiente]")),
        ("Inclusión", "Estudios incluidos en revisión final", counts.get("included_final", "[pendiente]")),
    ]
    return pd.DataFrame(rows, columns=["etapa_prisma", "descripcion", "n"])


def create_flow_text(counts: Dict) -> str:
    eric = counts.get("read_ERIC", "[n]")
    scopus = counts.get("read_Scopus", "[n]")
    wos = counts.get("read_WoS", "[n]")
    total = counts.get("read_total", "[n]")
    dup = counts.get("duplicates_removed", "[n]")
    unique = counts.get("unique_after_dedup", "[n]")
    screened = counts.get("screened_title_abstract", unique if unique != "[pendiente]" else "[n]")
    excluded = counts.get("excluded_title_abstract", "[pendiente]")
    full = counts.get("full_text_assessed", "[pendiente]")
    full_ex = counts.get("full_text_excluded", "[pendiente]")
    inc = counts.get("included_final", "[pendiente]")
    return f"""IDENTIFICACIÓN
ERIC = {eric} | Scopus = {scopus} | WoS = {wos}
Total exportado/leído = {total}
↓
DEDUPLICACIÓN
Duplicados eliminados = {dup}
Registros únicos = {unique}
↓
CRIBADO
Título y resumen evaluados = {screened}
Excluidos = {excluded}
↓
ELEGIBILIDAD
Textos completos evaluados = {full}
Excluidos con razones = {full_ex}
↓
INCLUSIÓN
Estudios incluidos = {inc}
"""


def create_flow_png(flow_text: str, out_path: Path) -> Path:
    out = versioned_path(out_path)
    fig = plt.figure(figsize=(8, 10))
    ax = fig.add_subplot(111)
    ax.axis("off")
    ax.text(0.5, 0.5, flow_text, ha="center", va="center", fontsize=13,
            bbox=dict(boxstyle="round,pad=0.8", fc="white", ec="black", lw=1.2))
    fig.tight_layout()
    fig.savefig(out, dpi=200, bbox_inches="tight")
    plt.close(fig)
    logging.info("Flujo PRISMA PNG exportado: %s", out)
    return out


def update_prisma_outputs(dirs: Dict[str, Path], counts: Dict) -> None:
    table = build_prisma_table(counts)
    write_excel(dirs["07_prisma_outputs"] / "tabla_prisma.xlsx", {"tabla_prisma": table})
    flow_text = create_flow_text(counts)
    flow_path = versioned_path(dirs["07_prisma_outputs"] / "flujo_prisma_ppt.txt")
    flow_path.write_text(flow_text, encoding="utf-8")
    logging.info("Flujo PRISMA texto exportado: %s", flow_path)
    create_flow_png(flow_text, dirs["07_prisma_outputs"] / "flujo_prisma.png")

# -----------------------------------------------------------------------------
# Modos principales
# -----------------------------------------------------------------------------
def prepare(root: Path, files: Dict[str, Path]) -> None:
    dirs = ensure_dirs(root)
    setup_logging(dirs["09_logs"], "prepare")
    logging.info("Root: %s", root)
    logging.info("rapidfuzz disponible: %s", HAS_RAPIDFUZZ)

    # respaldo de archivos
    for base, src in files.items():
        safe_copy(src, dirs["01_raw_exports"])

    diagnostics = []
    normalized_frames = []
    read_counts = {}
    all_warnings = []

    for base, path in files.items():
        if not path.exists():
            logging.error("Archivo no encontrado para %s: %s", base, path)
            continue
        df_raw, enc, sep, warns = read_csv_robust(path)
        all_warnings.extend([{"base": base, "advertencia": w} for w in warns])
        diag = diagnostic_for_df(df_raw, base, path, enc, sep)
        diagnostics.append(diag)
        read_counts[base] = len(df_raw)
        norm = normalize_base(df_raw, base)
        normalized_frames.append(norm)
        expected = EXPECTED_COUNTS.get(base)
        if expected is not None and expected != len(df_raw):
            msg = f"Diferencia de conteo {base}: esperado={expected}, leído={len(df_raw)}"
            logging.warning(msg)
            all_warnings.append({"base": base, "advertencia": msg})

    if not normalized_frames:
        raise RuntimeError("No se leyó ningún archivo. Revisa rutas.")

    diag_df = pd.DataFrame(diagnostics)
    expected_total = sum(EXPECTED_COUNTS.values())
    read_total = sum(read_counts.values())
    if expected_total != read_total:
        msg = f"Diferencia total: esperado={expected_total}, leído={read_total}. Se procesan registros leídos."
        logging.warning(msg)
        all_warnings.append({"base": "TOTAL", "advertencia": msg})

    warnings_df = pd.DataFrame(all_warnings) if all_warnings else pd.DataFrame(columns=["base", "advertencia"])
    write_excel(dirs["02_diagnostics"] / "diagnostico_archivos.xlsx", {"diagnostico": diag_df, "advertencias": warnings_df})
    txt_diag = dirs["02_diagnostics"] / "diagnostico_archivos.txt"
    txt_diag.write_text(diag_df.to_string(index=False) + "\n\nADVERTENCIAS\n" + warnings_df.to_string(index=False), encoding="utf-8")
    logging.info("Diagnóstico texto exportado: %s", txt_diag)

    unified = pd.concat(normalized_frames, ignore_index=True)
    # Rehacer id_unico global para evitar colisiones y mantener base.
    unified["id_unico"] = [f"REG_{i+1:05d}" for i in range(len(unified))]

    resumen_por_base = unified.groupby("base_origen").agg(
        registros=("id_unico", "count"),
        titulos_vacios=("titulo", lambda s: int(s.fillna("").astype(str).str.strip().eq("").sum())),
        abstracts_vacios=("abstract", lambda s: int(s.fillna("").astype(str).str.strip().eq("").sum())),
        doi_vacios=("doi_limpio", lambda s: int(s.fillna("").astype(str).str.strip().eq("").sum())),
        anios_invalidos=("anio_valido_2021_2026", lambda s: int((s != "sí").sum())),
        filas_anomalas=("fila_anomala_si_no", lambda s: int((s == "sí").sum())),
    ).reset_index()
    resumen_por_base["registros_esperados"] = resumen_por_base["base_origen"].map(EXPECTED_COUNTS)
    resumen_por_base["diferencia"] = resumen_por_base["registros"] - resumen_por_base["registros_esperados"]

    write_excel(dirs["03_unified_database"] / "base_unificada.xlsx", {
        "base_unificada": unified,
        "resumen_por_base": resumen_por_base,
        "advertencias": warnings_df,
    })
    unified.to_csv(versioned_path(dirs["03_unified_database"] / "base_unificada.csv"), index=False, encoding="utf-8-sig")

    marked, dedup, probable_df, dup_summary = mark_duplicates(unified)
    groups_df = dup_summary.pop("groups_df")
    summary_df = pd.DataFrame([dup_summary])
    write_excel(dirs["04_deduplication"] / "base_con_duplicados_marcados.xlsx", {"duplicados_marcados": marked})
    write_excel(dirs["04_deduplication"] / "base_deduplicada.xlsx", {"base_deduplicada": dedup})
    write_excel(dirs["04_deduplication"] / "reporte_duplicados.xlsx", {
        "resumen": summary_df,
        "grupos_duplicados": groups_df,
        "duplicados_probables": probable_df,
    })

    screening = add_screening_suggestions(dedup)
    screening_cols = [
        "id_unico", "base_origen_combinada", "titulo", "abstract", "autores", "año", "revista_fuente", "doi", "idioma",
        "tipo_documento", "palabras_clave", "url", "texto_screening", "sugerencia_auto_titulo_resumen",
        "razon_sugerencia_auto", "criterios_detectados", "advertencia_revision_manual", "decision_titulo_resumen",
        "razon_exclusion_titulo_resumen", "notas_screening",
    ]
    for c in screening_cols:
        if c not in screening.columns:
            screening[c] = ""
    screen_path = write_excel(dirs["05_title_abstract_screening"] / "matriz_screening_titulo_resumen.xlsx", {
        "matriz_screening": screening[screening_cols],
        "instrucciones_screening": screening_instructions_df(),
    })
    add_validations_to_screening(screen_path)

    resumen_auto = {
        "conteo_sugerencias": screening["sugerencia_auto_titulo_resumen"].value_counts(dropna=False).rename_axis("sugerencia").reset_index(name="n"),
        "conteo_razones": screening["razon_sugerencia_auto"].value_counts(dropna=False).rename_axis("razon").reset_index(name="n"),
        "conteo_por_base": screening["base_origen_combinada"].value_counts(dropna=False).rename_axis("base_origen_combinada").reset_index(name="n"),
        "registros_sin_abstract": screening[screening["tiene_abstract"] == "no"] if "tiene_abstract" in screening.columns else pd.DataFrame(),
        "registros_anio_invalido": screening[screening["anio_valido_2021_2026"] != "sí"] if "anio_valido_2021_2026" in screening.columns else pd.DataFrame(),
        "registros_sin_doi": screening[screening["tiene_doi"] == "no"] if "tiene_doi" in screening.columns else pd.DataFrame(),
        "registros_anomalos": screening[screening["fila_anomala_si_no"] == "sí"] if "fila_anomala_si_no" in screening.columns else pd.DataFrame(),
        "duplicados_probables": probable_df,
    }
    write_excel(dirs["05_title_abstract_screening"] / "resumen_screening_auto.xlsx", resumen_auto)

    create_full_text_template(dedup, screen_path, dirs["06_full_text_screening"] / "matriz_texto_completo.xlsx")
    create_extraction_template(dirs["08_final_corpus"] / "matriz_extraccion_final.xlsx")

    counts = {
        "expected_ERIC": EXPECTED_COUNTS["ERIC"],
        "expected_Scopus": EXPECTED_COUNTS["Scopus"],
        "expected_WoS": EXPECTED_COUNTS["WoS"],
        "expected_total": expected_total,
        "read_ERIC": read_counts.get("ERIC", 0),
        "read_Scopus": read_counts.get("Scopus", 0),
        "read_WoS": read_counts.get("WoS", 0),
        "read_total": read_total,
        "duplicates_removed": dup_summary["duplicados_eliminados"],
        "unique_after_dedup": dup_summary["registros_unicos_finales"],
        "screened_title_abstract": dup_summary["registros_unicos_finales"],
    }
    update_prisma_outputs(dirs, counts)

    # Resumen metodológico
    warn_text = "sí" if expected_total != read_total or any(v != EXPECTED_COUNTS.get(k, v) for k, v in read_counts.items()) else "no"
    method_text = f"""Tras aplicar filtros iniciales en las bases de datos, se procesaron los archivos exportados desde ERIC, Scopus y Web of Science mediante un pipeline reproducible en Python. La investigadora esperaba 581 registros exportados (ERIC=118, Scopus=223, WoS=240); el script leyó {read_total} registros (ERIC={read_counts.get('ERIC', 0)}, Scopus={read_counts.get('Scopus', 0)}, WoS={read_counts.get('WoS', 0)}). Las diferencias de conteo fueron registradas como advertencias y se procesaron los registros realmente leídos. El procedimiento permitió diagnosticar los archivos, marcar anomalías, normalizar campos bibliográficos, unificar registros y realizar deduplicación por DOI limpio, título normalizado y similitud de título. Se eliminaron {dup_summary['duplicados_eliminados']} duplicados exactos y quedaron {dup_summary['registros_unicos_finales']} registros únicos para cribado por título y resumen. Las sugerencias automáticas generadas por el script se utilizaron solo como apoyo organizativo; la decisión final de inclusión o exclusión corresponde a la investigadora de acuerdo con criterios previamente definidos. El proceso quedó documentado en una tabla y flujo tipo PRISMA.\n"""
    (dirs["09_logs"] / "resumen_metodologico.txt").write_text(method_text, encoding="utf-8")

    # respaldo del script
    try:
        current_script = Path(__file__).resolve()
        safe_copy(current_script, dirs["10_scripts_backup"])
    except Exception:
        pass

    print("\nProceso completado.")
    print(f"ERIC leído: {read_counts.get('ERIC', 0)}.")
    print(f"Scopus leído: {read_counts.get('Scopus', 0)}.")
    print(f"WoS leído: {read_counts.get('WoS', 0)}.")
    print(f"Total leído: {read_total}.")
    print(f"Advertencias de conteo: {warn_text}.")
    print(f"Duplicados eliminados: {dup_summary['duplicados_eliminados']}.")
    print(f"Registros únicos para cribado: {dup_summary['registros_unicos_finales']}.")
    print(f"Duplicados probables para revisión manual: {dup_summary['duplicados_probables_revision_manual']}.")
    print(f"Registros con año inválido: {int((unified['anio_valido_2021_2026'] != 'sí').sum())}.")
    print(f"Registros con filas anómalas: {int((unified['fila_anomala_si_no'] == 'sí').sum())}.")
    print("Revisar manualmente:")
    print(dirs["05_title_abstract_screening"] / "matriz_screening_titulo_resumen.xlsx")
    print("La decisión final de inclusión/exclusión debe ser realizada por la investigadora.")


def update_screening(root: Path) -> None:
    dirs = ensure_dirs(root)
    setup_logging(dirs["09_logs"], "update_screening")
    screen_path = dirs["05_title_abstract_screening"] / "matriz_screening_titulo_resumen.xlsx"
    dedup_path = dirs["04_deduplication"] / "base_deduplicada.xlsx"
    if not screen_path.exists():
        raise FileNotFoundError(f"No existe matriz de screening: {screen_path}")
    screen = pd.read_excel(screen_path, sheet_name="matriz_screening")
    decisions = screen["decision_titulo_resumen"].fillna("").astype(str).str.strip().str.lower() if "decision_titulo_resumen" in screen.columns else pd.Series([""] * len(screen))

    sent = decisions.isin(["incluir", "dudoso"])
    excluded = decisions.eq("excluir")
    pending = decisions.eq("")

    full_template = screen.loc[sent].copy()
    if full_template.empty:
        full_template = screen.copy()
        logging.warning("No hay decisiones incluir/dudoso; se crea plantilla de texto completo con todos los registros.")

    cols = ["id_unico", "titulo", "autores", "año", "revista_fuente", "doi", "url", "base_origen_combinada", "decision_titulo_resumen", "razon_exclusion_titulo_resumen"]
    for c in cols:
        if c not in full_template.columns:
            full_template[c] = ""
    out = full_template[cols].copy()
    for c in ["texto_completo_buscado_si_no", "texto_completo_recuperado_si_no", "decision_texto_completo", "razon_exclusion_texto_completo", "decision_final", "notas_texto_completo"]:
        if c not in out.columns:
            out[c] = ""
    write_excel(dirs["06_full_text_screening"] / "matriz_texto_completo.xlsx", {"matriz_texto_completo": out})

    # intentar recuperar conteos previos desde tabla_prisma
    counts = {"screened_title_abstract": len(screen), "excluded_title_abstract": int(excluded.sum()), "sent_full_text": int(sent.sum())}
    # Duplicación desde base dedup si existe
    if dedup_path.exists():
        dedup = pd.read_excel(dedup_path, sheet_name=0)
        counts["unique_after_dedup"] = len(dedup)
        counts["screened_title_abstract"] = len(dedup)
    update_prisma_outputs(dirs, counts)
    logging.info("Screening actualizado: incluir/dudoso=%s, excluir=%s, pendientes=%s", int(sent.sum()), int(excluded.sum()), int(pending.sum()))
    print("Actualización de screening completada.")
    print(f"Incluir/dudoso enviados a texto completo: {int(sent.sum())}")
    print(f"Excluidos por título/resumen: {int(excluded.sum())}")
    print(f"Pendientes sin decisión: {int(pending.sum())}")


def finalize(root: Path) -> None:
    dirs = ensure_dirs(root)
    setup_logging(dirs["09_logs"], "finalize")
    full_path = dirs["06_full_text_screening"] / "matriz_texto_completo.xlsx"
    if not full_path.exists():
        raise FileNotFoundError(f"No existe matriz de texto completo: {full_path}")
    full = pd.read_excel(full_path, sheet_name="matriz_texto_completo")

    def lower_col(name: str) -> pd.Series:
        if name not in full.columns:
            return pd.Series([""] * len(full))
        return full[name].fillna("").astype(str).str.strip().str.lower()

    buscado = lower_col("texto_completo_buscado_si_no").isin(["sí", "si", "yes", "1"])
    recuperado = lower_col("texto_completo_recuperado_si_no").isin(["sí", "si", "yes", "1"])
    decision_tc = lower_col("decision_texto_completo")
    decision_final = lower_col("decision_final")

    included_mask = decision_final.isin(["incluido", "incluir", "sí", "si", "included"]) | decision_tc.isin(["incluir", "incluido"])
    excluded_mask = decision_tc.isin(["excluir", "excluido"])
    not_retrieved = buscado & ~recuperado
    assessed = recuperado | decision_tc.isin(["incluir", "incluido", "excluir", "excluido"])

    corpus = full.loc[included_mask].copy()
    write_excel(dirs["08_final_corpus"] / "corpus_final.xlsx", {"corpus_final": corpus})
    create_extraction_template(dirs["08_final_corpus"] / "matriz_extraccion_final.xlsx", corpus)

    counts = {
        "full_text_not_retrieved": int(not_retrieved.sum()),
        "full_text_assessed": int(assessed.sum()),
        "full_text_excluded": int(excluded_mask.sum()),
        "included_final": int(included_mask.sum()),
    }
    update_prisma_outputs(dirs, counts)
    logging.info("Finalize completado: incluidos=%s", int(included_mask.sum()))
    print("Finalización completada.")
    print(f"Textos completos no recuperados: {int(not_retrieved.sum())}")
    print(f"Textos completos evaluados: {int(assessed.sum())}")
    print(f"Textos completos excluidos: {int(excluded_mask.sum())}")
    print(f"Estudios incluidos: {int(included_mask.sum())}")

# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pipeline PRISMA para ansiedad matemática en escolares")
    parser.add_argument("--mode", choices=["prepare", "update_screening", "finalize"], default="prepare")
    parser.add_argument("--root", type=str, default=str(DEFAULT_ROOT), help="Carpeta raíz del proyecto")
    parser.add_argument("--eric", type=str, default=str(DEFAULT_FILES["ERIC"]), help="Ruta CSV ERIC")
    parser.add_argument("--scopus", type=str, default=str(DEFAULT_FILES["Scopus"]), help="Ruta CSV Scopus")
    parser.add_argument("--wos", type=str, default=str(DEFAULT_FILES["WoS"]), help="Ruta CSV Web of Science")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root)
    files = {"ERIC": Path(args.eric), "Scopus": Path(args.scopus), "WoS": Path(args.wos)}
    if args.mode == "prepare":
        prepare(root, files)
    elif args.mode == "update_screening":
        update_screening(root)
    elif args.mode == "finalize":
        finalize(root)


if __name__ == "__main__":
    main()

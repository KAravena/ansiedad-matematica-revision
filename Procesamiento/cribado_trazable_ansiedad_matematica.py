# -*- coding: utf-8 -*-
"""
Cribado asistido trazable para revisión sistemática exploratoria PRISMA.
Tema: Ansiedad matemática en estudiantes escolares.

Este script NO decide la inclusión final de artículos. Aplica reglas explícitas
sobre título, resumen y palabras clave para generar una preclasificación
asistida: incluir_probable, excluir_probable o dudoso_revision_humana.
Luego prioriza solo los registros no excluidos y exporta 12 candidatos
preliminares para validación humana.

Uso básico en Windows:
    python cribado_trazable_ansiedad_matematica.py

Uso con matriz explícita:
    python cribado_trazable_ansiedad_matematica.py --input "D:\\Psicologia\\Seminario de investigación\\presentación\\05_title_abstract_screening\\matriz_screening_titulo_resumen.xlsx"

Uso con carpeta raíz distinta:
    python cribado_trazable_ansiedad_matematica.py --root "D:\\Psicologia\\Seminario de investigación\\presentación"

Requisitos:
    pandas, openpyxl, unidecode
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from unidecode import unidecode
except Exception:  # pragma: no cover
    def unidecode(x: str) -> str:
        return x

# -----------------------------------------------------------------------------
# Configuración general
# -----------------------------------------------------------------------------
DEFAULT_ROOT = Path(r"D:\Psicologia\Seminario de investigación\presentación")
DEFAULT_INPUT = DEFAULT_ROOT / "05_title_abstract_screening" / "matriz_screening_titulo_resumen.xlsx"

# Cifras ya producidas por el pipeline previo; se usan solo para reporte.
N_REGISTROS_LEIDOS = 591
N_DUPLICADOS_ELIMINADOS = 245

OUTPUT_DIRS = [
    "05_title_abstract_screening",
    "07_prisma_outputs",
    "08_final_corpus",
    "09_logs",
]

BASE_COLUMNS = [
    "id_unico", "base_origen_combinada", "titulo", "abstract", "autores", "año",
    "revista_fuente", "doi", "idioma", "tipo_documento", "palabras_clave", "url",
    "texto_screening", "sugerencia_auto_titulo_resumen", "razon_sugerencia_auto",
    "decision_titulo_resumen", "razon_exclusion_titulo_resumen", "notas_screening",
]

# -----------------------------------------------------------------------------
# Patrones de inclusión/exclusión
# -----------------------------------------------------------------------------
INCLUSION_PATTERNS: Dict[str, List[str]] = {
    "I1_estudio_empirico": [
        r"\bsample\b", r"\bparticipants?\b", r"\bstudents?\b", r"\bdata\b",
        r"\bsurvey\b", r"\bquestionnaire\b", r"\bscale\b", r"\bmeasured\b",
        r"\banalysis\b", r"\bregression\b", r"\bmodel\b", r"\bexperiment\b",
        r"\bcross-sectional\b", r"\blongitudinal\b", r"\bpisa\b", r"\btimss\b",
        r"\bsem\b", r"\bmultilevel\b",
    ],
    "I3_poblacion_escolar": [
        r"\bschools?\b", r"\bstudents?\b", r"\bpupils?\b", r"\bchild\b", r"\bchildren\b",
        r"\badolescents?\b", r"\bprimary\b", r"\bsecondary\b", r"\belementary\b",
        r"\bmiddle school\b", r"\bhigh school\b", r"\bgrade\b", r"\bgrades?\s*\d+\b",
        r"\bclassroom\b", r"\bschool-aged\b",
    ],
    "I4_ansiedad_matematica_explicita": [
        r"\bmath anxiety\b", r"\bmathematics anxiety\b", r"\bmathematical anxiety\b",
        r"\bansiedad matematica\b", r"\bansiedade matematica\b", r"\bmath-anxiety\b",
    ],
    "I5_dimension_educativa": [
        r"\bachievement\b", r"\bperformance\b", r"\blearning\b", r"\bengagement\b",
        r"\bstrateg(?:y|ies)\b", r"\bgender\b", r"\bsocioeconomic\b", r"\bses\b",
        r"\bschool context\b", r"\bclassroom\b", r"\battainment\b", r"\bmathematical literacy\b",
        r"\bself-efficacy\b", r"\battitudes?\b", r"\bmotivation\b", r"\bworking memory\b",
        r"\bavoidance\b", r"\bteacher support\b", r"\bfamily support\b", r"\bparental math anxiety\b",
    ],
}

EXCLUSION_PATTERNS: Dict[str, List[str]] = {
    "E1_revision_meta": [
        r"\bsystematic review\b", r"\bmeta-analysis\b", r"\bmeta analysis\b",
        r"\bliterature review\b", r"\bscoping review\b", r"\bnarrative review\b",
        r"\breview article\b",
    ],
    "E2_educacion_superior": [
        r"\buniversity students?\b", r"\bundergraduate\b", r"\bcollege students?\b",
        r"\bhigher education\b", r"\btertiary education\b",
    ],
    "E3_formacion_docente": [
        r"\bpreservice teacher\b", r"\bpre-service teacher\b", r"\bteacher education\b",
        r"\bteacher candidates?\b", r"\bteachers only\b", r"\bin-service teachers?\b",
    ],
    "E4_no_articulo_no_empirico": [
        r"\beditorial\b", r"\bcommentary\b", r"\bbook chapter\b", r"\bdissertation\b",
        r"\bthesis\b", r"\breport\b", r"\bconference abstract\b",
    ],
    "E6_poblacion_no_escolar_clara": [
        r"\badults?\b", r"\buniversity\b", r"\bundergraduate\b", r"\bcollege\b",
        r"\bteachers?\b", r"\bparents only\b",
    ],
}

MEASUREMENT_PATTERNS = [
    r"\bscale\b", r"\bquestionnaire\b", r"\binstrument\b", r"\bmeasure\b",
    r"\bmeasured\b", r"\bassessment\b", r"\bsurvey\b", r"\bpisa\b", r"\btimss\b",
    r"\bmamas\b", r"\bmasc\b", r"\bmath anxiety scale\b", r"\bmathematics anxiety scale\b",
]

METHOD_HIGH_PATTERNS = [
    r"\bstructural equation\b", r"\bsem\b", r"\bmultilevel\b", r"\bhierarchical\b",
    r"\blongitudinal\b", r"\bmediation\b", r"\bmediating\b", r"\blatent profile\b",
    r"\bpath analysis\b", r"\bpisa\b", r"\btimss\b", r"\blarge-scale\b", r"\bpanel\b",
]

METHOD_BASIC_PATTERNS = [
    r"\bcross-sectional\b", r"\bregression\b", r"\bsurvey\b", r"\bquantitative\b",
    r"\bcorrelation\b", r"\bcorrelational\b", r"\banova\b", r"\bmodel\b",
]

DIMENSION_GROUPS: Dict[str, List[str]] = {
    "rendimiento": [r"\bachievement\b", r"\bperformance\b", r"\battainment\b", r"\bmathematical literacy\b", r"\btest scores?\b"],
    "engagement": [r"\bengagement\b", r"\bparticipation\b"],
    "estrategias": [r"\bstrategy\b", r"\bstrategies\b", r"\bproblem solving\b"],
    "genero": [r"\bgender\b", r"\bgirls\b", r"\bboys\b", r"\bsex differences\b"],
    "nse": [r"\bsocioeconomic\b", r"\bses\b", r"\bsocio-economic\b", r"\bdisadvantaged\b"],
    "contexto": [r"\bschool context\b", r"\bclassroom\b", r"\bteacher support\b", r"\bschool climate\b", r"\bcontextual\b"],
    "actitudes": [r"\bself-efficacy\b", r"\battitudes?\b", r"\bmotivation\b", r"\bconfidence\b", r"\bmindset\b"],
    "cognitivo": [r"\bworking memory\b", r"\bavoidance\b", r"\bcognitive\b"],
}

SCHOOL_STRONG = [
    r"\bprimary\b", r"\bsecondary\b", r"\belementary\b", r"\bmiddle school\b",
    r"\bhigh school\b", r"\bgrade\s*\d+\b", r"\bgrades?\b", r"\bpupils?\b",
    r"\bchildren\b", r"\bchild\b", r"\badolescents?\b", r"\bschool-aged\b",
]

SCHOOL_GENERAL = [r"\bschool\b", r"\bschools\b", r"\bstudent\b", r"\bstudents\b", r"\bclassroom\b"]

CONTEXT_DIVERSITY = [
    r"\bunited states\b", r"\busa\b", r"\bchina\b", r"\bgermany\b", r"\baustralia\b",
    r"\buk\b", r"\bengland\b", r"\bpoland\b", r"\bswitzerland\b", r"\binternational\b",
    r"\bcross-national\b", r"\bcross cultural\b", r"\bcross-cultural\b", r"\bcountries\b", r"\bcountry\b",
]

# -----------------------------------------------------------------------------
# Utilidades generales
# -----------------------------------------------------------------------------
def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    txt = unidecode(str(value))
    txt = txt.lower()
    txt = re.sub(r"[\r\n\t]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def contains_any(text: str, patterns: Iterable[str]) -> bool:
    return any(re.search(p, text, flags=re.IGNORECASE) for p in patterns)


def matched_codes(text: str, patterns: Dict[str, List[str]]) -> List[str]:
    return [code for code, pats in patterns.items() if contains_any(text, pats)]


def matched_terms(text: str, patterns: Iterable[str]) -> List[str]:
    return [p for p in patterns if re.search(p, text, flags=re.IGNORECASE)]


def count_matches(text: str, patterns: Iterable[str]) -> int:
    return sum(1 for p in patterns if re.search(p, text, flags=re.IGNORECASE))


def versioned_path(path: Path) -> Path:
    if not path.exists():
        return path
    return path.with_name(f"{path.stem}_{now_tag()}{path.suffix}")


def ensure_output_dirs(root: Path) -> Dict[str, Path]:
    out = {name: root / name for name in OUTPUT_DIRS}
    for path in out.values():
        path.mkdir(parents=True, exist_ok=True)
    return out


def autosize_excel(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in list(col)[:200]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def read_screening_matrix(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"No existe la matriz de cribado: {path}")
    xls = pd.ExcelFile(path)
    sheet = "matriz_screening" if "matriz_screening" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet)
    for col in BASE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def build_screening_text(row: pd.Series) -> str:
    parts = [
        row.get("titulo", ""),
        row.get("abstract", ""),
        row.get("palabras_clave", ""),
        row.get("tipo_documento", ""),
        row.get("revista_fuente", ""),
        row.get("texto_screening", ""),
    ]
    return normalize_text(" ".join(str(x) for x in parts if not pd.isna(x)))


def abstract_is_sufficient(row: pd.Series, min_chars: int = 300) -> bool:
    return len(normalize_text(row.get("abstract", ""))) >= min_chars


def valid_year(row: pd.Series) -> bool:
    raw = row.get("año", "")
    try:
        y = int(float(str(raw).strip()))
        return 2021 <= y <= 2026
    except Exception:
        return False

# -----------------------------------------------------------------------------
# Clasificación asistida por reglas
# -----------------------------------------------------------------------------
def detect_inclusion_criteria(row: pd.Series, text: str) -> Dict[str, bool]:
    detected = {
        "I1_estudio_empirico": contains_any(text, INCLUSION_PATTERNS["I1_estudio_empirico"]),
        "I2_periodo_2021_2026": valid_year(row),
        "I3_poblacion_escolar": contains_any(text, INCLUSION_PATTERNS["I3_poblacion_escolar"]),
        "I4_ansiedad_matematica_explicita": contains_any(text, INCLUSION_PATTERNS["I4_ansiedad_matematica_explicita"]),
        "I5_dimension_educativa": contains_any(text, INCLUSION_PATTERNS["I5_dimension_educativa"]),
    }
    return detected


def detect_exclusion_criteria(row: pd.Series, text: str, inc: Dict[str, bool]) -> Dict[str, bool]:
    exc = {code: contains_any(text, pats) for code, pats in EXCLUSION_PATTERNS.items()}

    # E5: ausencia de ansiedad matemática explícita. Si el abstract es insuficiente,
    # no se excluye automáticamente; se manda a revisión humana.
    if not inc["I4_ansiedad_matematica_explicita"] and abstract_is_sufficient(row):
        exc["E5_no_ansiedad_matematica_explicita"] = True
    else:
        exc["E5_no_ansiedad_matematica_explicita"] = False

    # E6 solo opera si hay población no escolar clara y no hay señales escolares.
    if exc.get("E6_poblacion_no_escolar_clara", False) and inc["I3_poblacion_escolar"]:
        exc["E6_poblacion_no_escolar_clara"] = False

    return exc


def classify_record(row: pd.Series) -> Dict[str, object]:
    text = build_screening_text(row)
    inc = detect_inclusion_criteria(row, text)
    exc = detect_exclusion_criteria(row, text, inc)

    inc_codes = [k for k, v in inc.items() if v]
    exc_codes = [k for k, v in exc.items() if v]

    year_ok = inc["I2_periodo_2021_2026"]
    abstract_ok = abstract_is_sufficient(row)
    strong_exc = [c for c in exc_codes if c in ["E1_revision_meta", "E2_educacion_superior", "E3_formacion_docente", "E4_no_articulo_no_empirico"]]
    no_math_exc = exc.get("E5_no_ansiedad_matematica_explicita", False)
    non_school_exc = exc.get("E6_poblacion_no_escolar_clara", False)

    requires_human = "no"
    confidence = "media"
    reason = ""

    # Reglas de seguridad: año inválido o abstract insuficiente = revisión humana,
    # salvo que la exclusión fuerte sea muy clara. Para evitar falsos positivos, se prioriza duda.
    if not year_ok:
        classification = "dudoso_revision_humana"
        requires_human = "sí"
        confidence = "baja"
        reason = "año inválido o fuera de rango; requiere verificación"
    elif not abstract_ok and not strong_exc:
        classification = "dudoso_revision_humana"
        requires_human = "sí"
        confidence = "baja"
        reason = "sin abstract suficiente para decidir"
    elif strong_exc:
        classification = "excluir_probable"
        requires_human = "no"
        confidence = "alta"
        reason = "exclusión fuerte detectada: " + "; ".join(strong_exc)
    elif non_school_exc:
        classification = "excluir_probable"
        requires_human = "no"
        confidence = "alta"
        reason = "población no escolar clara sin señales escolares"
    elif no_math_exc:
        classification = "excluir_probable"
        requires_human = "no"
        confidence = "alta"
        reason = "no mide ansiedad matemática explícita"
    elif all(inc.values()):
        classification = "incluir_probable"
        requires_human = "sí"  # aun así debe validarse; no es inclusión final
        confidence = "alta"
        reason = "cumple criterios asistidos I1-I5: empírico + periodo + escolar + ansiedad matemática + dimensión educativa"
    elif inc["I2_periodo_2021_2026"] and inc["I3_poblacion_escolar"] and inc["I4_ansiedad_matematica_explicita"] and inc["I5_dimension_educativa"]:
        classification = "dudoso_revision_humana"
        requires_human = "sí"
        confidence = "media"
        reason = "cumple escolar + ansiedad matemática + dimensión educativa, pero falta evidencia textual clara de estudio empírico"
    else:
        classification = "dudoso_revision_humana"
        requires_human = "sí"
        confidence = "baja"
        missing = [k for k, v in inc.items() if not v]
        reason = "evidencia insuficiente; criterios no detectados: " + "; ".join(missing)

    return {
        "texto_cribado_normalizado": text,
        "clasificacion_asistida": classification,
        "razon_clasificacion_asistida": reason,
        "criterios_inclusion_detectados": "; ".join(inc_codes),
        "criterios_exclusion_detectados": "; ".join(exc_codes),
        "requiere_revision_humana": requires_human,
        "nivel_confianza_asistida": confidence,
        "abstract_suficiente": "sí" if abstract_ok else "no",
        "anio_valido_2021_2026": "sí" if year_ok else "no",
    }

# -----------------------------------------------------------------------------
# Rúbrica 0-14 aplicada a incluir_probable/dudoso
# -----------------------------------------------------------------------------
def matched_dimension_labels(text: str) -> List[str]:
    return [label for label, pats in DIMENSION_GROUPS.items() if contains_any(text, pats)]


def score_record(row: pd.Series) -> Dict[str, object]:
    text = row.get("texto_cribado_normalizado", "") or build_screening_text(row)
    title = normalize_text(row.get("titulo", ""))

    # No se priorizan excluidos; quedan con puntaje 0 para evitar que entren al top 12.
    if row.get("clasificacion_asistida", "") == "excluir_probable":
        return {
            "C1_centralidad_ansiedad_matematica": 0,
            "C2_poblacion_escolar": 0,
            "C3_medicion_ansiedad": 0,
            "C4_dimension_educativa": 0,
            "C5_aporte_metodologico": 0,
            "C6_aporte_objetivos_revision": 0,
            "C7_diversidad_corpus": 0,
            "puntaje_prioridad_0_14": 0,
            "nivel_prioridad": "excluido_probable",
            "justificacion_prioridad": "No se prioriza porque fue clasificado como excluir_probable.",
            "dimensiones_detectadas": "",
        }

    # C1 Centralidad de ansiedad matemática
    anxiety_count = count_matches(text, INCLUSION_PATTERNS["I4_ansiedad_matematica_explicita"])
    anxiety_title = contains_any(title, INCLUSION_PATTERNS["I4_ansiedad_matematica_explicita"])
    c1 = 2 if (anxiety_title or anxiety_count >= 2) else (1 if anxiety_count == 1 else 0)

    # C2 Población escolar clara
    school_strong = contains_any(text, SCHOOL_STRONG)
    school_general = contains_any(text, SCHOOL_GENERAL)
    c2 = 2 if (school_strong and school_general) else (1 if (school_strong or school_general) else 0)

    # C3 Medición explícita
    measurement_count = count_matches(text, MEASUREMENT_PATTERNS)
    c3 = 2 if measurement_count >= 2 or contains_any(text, [r"\bpisa\b", r"\btimss\b", r"\bmamas\b", r"\bmasc\b", r"\bmath anxiety scale\b"]) else (1 if measurement_count == 1 else 0)

    # C4 Dimensiones educativas
    dim_labels = matched_dimension_labels(text)
    c4 = 2 if len(dim_labels) >= 2 else (1 if len(dim_labels) == 1 else 0)

    # C5 Aporte metodológico
    c5 = 2 if contains_any(text, METHOD_HIGH_PATTERNS) else (1 if contains_any(text, METHOD_BASIC_PATTERNS) else 0)

    # C6 Aporte a objetivos
    objectives = 0
    if c1 >= 1 and c3 >= 1:
        objectives += 1  # definición/operacionalización
    if c5 >= 1:
        objectives += 1  # diseño/método
    if c4 >= 1:
        objectives += 1  # dimensiones asociadas
    c6 = 2 if objectives >= 2 else (1 if objectives == 1 else 0)

    # C7 Diversidad del corpus, aproximada por señales textuales
    diversity = 0
    if contains_any(text, CONTEXT_DIVERSITY):
        diversity += 1
    if contains_any(text, [r"\blongitudinal\b", r"\bmultilevel\b", r"\blatent profile\b", r"\bsem\b"]):
        diversity += 1
    if contains_any(text, [r"\bpisa\b", r"\btimss\b", r"\bmamas\b", r"\bmasc\b"]):
        diversity += 1
    if len(dim_labels) >= 3:
        diversity += 1
    c7 = 2 if diversity >= 2 else (1 if diversity == 1 else 0)

    total = c1 + c2 + c3 + c4 + c5 + c6 + c7
    if total >= 11:
        level = "alta"
    elif total >= 8:
        level = "media"
    elif total >= 5:
        level = "revisar"
    else:
        level = "baja"

    justification = (
        f"C1={c1}, C2={c2}, C3={c3}, C4={c4}, C5={c5}, C6={c6}, C7={c7}. "
        f"Dimensiones detectadas: {', '.join(dim_labels) if dim_labels else 'no claras'}."
    )

    return {
        "C1_centralidad_ansiedad_matematica": c1,
        "C2_poblacion_escolar": c2,
        "C3_medicion_ansiedad": c3,
        "C4_dimension_educativa": c4,
        "C5_aporte_metodologico": c5,
        "C6_aporte_objetivos_revision": c6,
        "C7_diversidad_corpus": c7,
        "puntaje_prioridad_0_14": total,
        "nivel_prioridad": level,
        "justificacion_prioridad": justification,
        "dimensiones_detectadas": "; ".join(dim_labels),
    }


def apply_classification_and_scoring(df: pd.DataFrame) -> pd.DataFrame:
    classifications = df.apply(classify_record, axis=1, result_type="expand")
    out = pd.concat([df.copy(), classifications], axis=1)
    scores = out.apply(score_record, axis=1, result_type="expand")
    out = pd.concat([out, scores], axis=1)
    return out

# -----------------------------------------------------------------------------
# Selección top 12 trazable
# -----------------------------------------------------------------------------
def assign_top12(df: pd.DataFrame, top_n: int = 12) -> pd.DataFrame:
    out = df.copy()
    out["estatus_candidato_12"] = "no_candidato"
    out.loc[out["clasificacion_asistida"] == "excluir_probable", "estatus_candidato_12"] = "excluido_probable"
    out["razon_entrada_top12"] = "no entra por menor puntaje o por requerir revisión posterior"
    out.loc[out["clasificacion_asistida"] == "excluir_probable", "razon_entrada_top12"] = out.loc[out["clasificacion_asistida"] == "excluir_probable", "razon_clasificacion_asistida"]
    out["orden_top12"] = ""

    sort_cols = [
        "puntaje_prioridad_0_14",
        "C1_centralidad_ansiedad_matematica",
        "C2_poblacion_escolar",
        "C3_medicion_ansiedad",
        "C4_dimension_educativa",
        "C5_aporte_metodologico",
    ]
    ascending = [False, False, False, False, False, False]

    include_pool = out[out["clasificacion_asistida"] == "incluir_probable"].sort_values(sort_cols, ascending=ascending)
    selected = include_pool.head(top_n).index.tolist()

    if len(selected) < top_n:
        remaining = top_n - len(selected)
        dudoso_pool = out[(out["clasificacion_asistida"] == "dudoso_revision_humana") & (~out.index.isin(selected))].sort_values(sort_cols, ascending=ascending)
        selected += dudoso_pool.head(remaining).index.tolist()

    for order, idx in enumerate(selected, start=1):
        if out.loc[idx, "clasificacion_asistida"] == "incluir_probable":
            out.loc[idx, "estatus_candidato_12"] = "candidato_preliminar"
            out.loc[idx, "razon_entrada_top12"] = "alta prioridad relativa y cumple criterios asistidos de inclusión probable"
        else:
            out.loc[idx, "estatus_candidato_12"] = "candidato_condicional"
            out.loc[idx, "razon_entrada_top12"] = "candidato condicional: requiere validación humana por evidencia insuficiente o ambigua"
        out.loc[idx, "orden_top12"] = f"{order:02d}"

    # Columnas para validación humana
    for col in [
        "cumple_criterios_obligatorios_revision_humana",
        "texto_completo_revisado_si_no",
        "seleccion_final_si_no",
        "razon_exclusion_manual_si_corresponde",
        "justificacion_inclusion_final",
        "notas_investigadora",
    ]:
        if col not in out.columns:
            out[col] = ""

    return out

# -----------------------------------------------------------------------------
# Reportes y textos
# -----------------------------------------------------------------------------
def criteria_rules_df() -> pd.DataFrame:
    rows = [
        ["I1", "Estudio empírico", "Busca señales como sample, participants, data, survey, regression, PISA, TIMSS, SEM, multilevel."],
        ["I2", "Periodo 2021–2026", "Usa columna año. Año inválido se manda a revisión humana."],
        ["I3", "Población escolar", "Busca school, student, child, adolescent, primary, secondary, grade, classroom."],
        ["I4", "Ansiedad matemática explícita", "Busca math anxiety, mathematics anxiety, mathematical anxiety, ansiedad matematica."],
        ["I5", "Dimensión educativa", "Busca achievement, performance, learning, engagement, gender, SES, classroom, motivation, etc."],
        ["E1", "Revisión/metaanálisis", "systematic review, meta-analysis, literature review, scoping review."],
        ["E2", "Educación superior", "university students, undergraduate, college students, higher education."],
        ["E3", "Formación docente", "pre-service teacher, teacher education, teachers only."],
        ["E4", "Documento no artículo/no empírico", "editorial, commentary, book chapter, dissertation, thesis, report."],
        ["E5", "No ansiedad matemática explícita", "Se aplica si no hay términos de ansiedad matemática y el abstract es suficiente."],
        ["E6", "Población no escolar clara", "adults, university, teachers, parents only sin señales escolares."],
        ["Regla", "No inclusión final automática", "Todas las decisiones deben validarse humanamente."],
    ]
    return pd.DataFrame(rows, columns=["Código", "Criterio", "Regla operacional"])


def guide_validation_df() -> pd.DataFrame:
    rows = [
        ["1", "Abrir top12_trazable_para_revision.xlsx."],
        ["2", "Revisar cada candidato por título y resumen completo."],
        ["3", "Confirmar criterios obligatorios: empírico, 2021–2026, población escolar, ansiedad matemática explícita y dimensión educativa."],
        ["4", "Buscar y revisar texto completo cuando haya duda."],
        ["5", "Completar cumple_criterios_obligatorios_revision_humana."],
        ["6", "Completar seleccion_final_si_no y justificar la inclusión o exclusión."],
        ["7", "Actualizar cifras PRISMA finales solo después de esta validación humana."],
    ]
    return pd.DataFrame(rows, columns=["Paso", "Acción"])


def methodological_text(unique_n: int, top_n: int) -> str:
    return (
        f"Después de la deduplicación quedaron {unique_n} registros únicos para cribado por título y resumen. "
        "Para apoyar esta etapa se aplicó un cribado asistido por reglas, basado en criterios explícitos de inclusión y exclusión. "
        "El script buscó señales textuales en título, resumen y palabras clave relacionadas con ansiedad matemática, población escolar, "
        "medición del constructo y dimensiones educativas. También marcó señales de exclusión, como revisiones, educación superior, "
        "formación docente o documentos no empíricos. Cada registro fue clasificado como incluir probable, excluir probable o dudoso para revisión humana. "
        f"Posteriormente, los registros no excluidos fueron ordenados mediante una rúbrica de prioridad de 0 a 14 puntos. Los {top_n} registros de mayor prioridad "
        "fueron exportados como candidatos preliminares para revisión de texto completo. Esta clasificación no constituye una inclusión final automática; "
        "todas las decisiones deben ser verificadas por la investigadora."
    )


def make_summary_tables(df: pd.DataFrame, top_n: int) -> Dict[str, pd.DataFrame]:
    unique_n = len(df)
    summary = pd.DataFrame({
        "Indicador": [
            "Registros leídos",
            "Duplicados eliminados",
            "Registros únicos para cribado",
            "Excluir_probable por reglas",
            "Incluir_probable por reglas",
            "Dudoso/revisión humana",
            f"Candidatos preliminares top {top_n}",
            "Registros que requieren revisión humana",
            "Registros con año inválido/fuera de rango",
            "Registros sin abstract suficiente",
            "Estudios incluidos finales",
        ],
        "n": [
            N_REGISTROS_LEIDOS,
            N_DUPLICADOS_ELIMINADOS,
            unique_n,
            int((df["clasificacion_asistida"] == "excluir_probable").sum()),
            int((df["clasificacion_asistida"] == "incluir_probable").sum()),
            int((df["clasificacion_asistida"] == "dudoso_revision_humana").sum()),
            int(df["estatus_candidato_12"].isin(["candidato_preliminar", "candidato_condicional"]).sum()),
            int((df["requiere_revision_humana"] == "sí").sum()),
            int((df["anio_valido_2021_2026"] == "no").sum()),
            int((df["abstract_suficiente"] == "no").sum()),
            "[pendiente de validación humana]",
        ],
        "Nota": [
            "Cifra del pipeline previo",
            "Cifra del pipeline previo",
            "Matriz deduplicada",
            "Cribado asistido; no final",
            "Cribado asistido; requiere validación",
            "Revisión humana necesaria",
            "Subcorpus preliminar",
            "Incluye dudosos y probables a validar",
            "No se excluye automáticamente",
            "No se excluye automáticamente si falta evidencia",
            "Completar después de revisión humana",
        ]
    })

    exclusions = (
        df[df["criterios_exclusion_detectados"].fillna("") != ""]
        .assign(criterio=lambda x: x["criterios_exclusion_detectados"].str.split("; "))
        .explode("criterio")
        .groupby("criterio", dropna=False).size().reset_index(name="n")
        .sort_values("n", ascending=False)
    )

    inclusions = (
        df[df["criterios_inclusion_detectados"].fillna("") != ""]
        .assign(criterio=lambda x: x["criterios_inclusion_detectados"].str.split("; "))
        .explode("criterio")
        .groupby("criterio", dropna=False).size().reset_index(name="n")
        .sort_values("n", ascending=False)
    )

    pending = df[df["requiere_revision_humana"] == "sí"].copy()
    note = pd.DataFrame({"nota_metodologica": [
        "Estas cifras corresponden a cribado asistido por reglas y no sustituyen el cribado humano final.",
        methodological_text(unique_n, top_n),
    ]})
    return {
        "resumen_prisma_asistido": summary,
        "exclusiones_por_razon": exclusions,
        "inclusiones_por_criterio": inclusions,
        "pendientes_revision_humana": pending,
        "nota_metodologica": note,
    }

# -----------------------------------------------------------------------------
# Escritura de outputs
# -----------------------------------------------------------------------------
def write_outputs(df: pd.DataFrame, dirs: Dict[str, Path], top_n: int) -> Dict[str, Path]:
    outputs: Dict[str, Path] = {}

    # 1. Matriz completa de cribado asistido
    matriz_path = versioned_path(dirs["05_title_abstract_screening"] / "matriz_cribado_asistido_trazable.xlsx")
    with pd.ExcelWriter(matriz_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="matriz_completa")
        df[df["clasificacion_asistida"] == "incluir_probable"].to_excel(writer, index=False, sheet_name="incluir_probable")
        df[df["clasificacion_asistida"] == "excluir_probable"].to_excel(writer, index=False, sheet_name="excluir_probable")
        df[df["clasificacion_asistida"] == "dudoso_revision_humana"].to_excel(writer, index=False, sheet_name="dudosos_revision_humana")
        top12 = df[df["estatus_candidato_12"].isin(["candidato_preliminar", "candidato_condicional"])].sort_values("orden_top12")
        top12.to_excel(writer, index=False, sheet_name="top12_candidatos")
        criteria_rules_df().to_excel(writer, index=False, sheet_name="criterios_y_reglas")
    autosize_excel(matriz_path)
    outputs["matriz_cribado_asistido_trazable"] = matriz_path

    # 2. Reporte PRISMA asistido preliminar
    report_path = versioned_path(dirs["07_prisma_outputs"] / "reporte_prisma_asistido_preliminar.xlsx")
    summary_tables = make_summary_tables(df, top_n)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        for sheet, table in summary_tables.items():
            safe_sheet = sheet[:31]
            table.to_excel(writer, index=False, sheet_name=safe_sheet)
    autosize_excel(report_path)
    outputs["reporte_prisma_asistido_preliminar"] = report_path

    # 3. Top 12 trazable para revisión humana
    top_path = versioned_path(dirs["08_final_corpus"] / "top12_trazable_para_revision.xlsx")
    top12 = df[df["estatus_candidato_12"].isin(["candidato_preliminar", "candidato_condicional"])].sort_values("orden_top12")
    top12_just = top12[[
        c for c in [
            "orden_top12", "estatus_candidato_12", "razon_entrada_top12", "clasificacion_asistida",
            "razon_clasificacion_asistida", "puntaje_prioridad_0_14", "nivel_prioridad", "titulo",
            "autores", "año", "revista_fuente", "doi", "abstract", "criterios_inclusion_detectados",
            "criterios_exclusion_detectados", "justificacion_prioridad", "cumple_criterios_obligatorios_revision_humana",
            "texto_completo_revisado_si_no", "seleccion_final_si_no", "razon_exclusion_manual_si_corresponde",
            "justificacion_inclusion_final", "notas_investigadora"
        ] if c in top12.columns
    ]]
    with pd.ExcelWriter(top_path, engine="openpyxl") as writer:
        top12.to_excel(writer, index=False, sheet_name="top12")
        top12_just.to_excel(writer, index=False, sheet_name="top12_con_justificacion")
        guide_validation_df().to_excel(writer, index=False, sheet_name="guia_validacion_humana")
    autosize_excel(top_path)
    outputs["top12_trazable_para_revision"] = top_path

    # 4. Texto metodológico
    text_path = versioned_path(dirs["07_prisma_outputs"] / "texto_metodologico_cribado_trazable.txt")
    text_path.write_text(methodological_text(len(df), top_n), encoding="utf-8")
    outputs["texto_metodologico_cribado_trazable"] = text_path

    return outputs


def write_log(root_dirs: Dict[str, Path], outputs: Dict[str, Path], df: pd.DataFrame, input_path: Path, top_n: int) -> Path:
    log_path = root_dirs["09_logs"] / "cribado_trazable_log.txt"
    with open(log_path, "a", encoding="utf-8") as f:
        f.write("=" * 80 + "\n")
        f.write(f"Fecha: {datetime.now().isoformat(timespec='seconds')}\n")
        f.write(f"Input: {input_path}\n")
        f.write(f"Registros únicos procesados: {len(df)}\n")
        f.write(f"Excluir probable por reglas: {(df['clasificacion_asistida'] == 'excluir_probable').sum()}\n")
        f.write(f"Incluir probable por reglas: {(df['clasificacion_asistida'] == 'incluir_probable').sum()}\n")
        f.write(f"Dudoso/revisión humana: {(df['clasificacion_asistida'] == 'dudoso_revision_humana').sum()}\n")
        f.write(f"Top {top_n} candidatos preliminares: {df['estatus_candidato_12'].isin(['candidato_preliminar', 'candidato_condicional']).sum()}\n")
        for name, path in outputs.items():
            f.write(f"Output {name}: {path}\n")
        f.write("Advertencia: cribado asistido por reglas; requiere validación humana.\n")
    return log_path

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Cribado asistido trazable para revisión sistemática exploratoria.")
    parser.add_argument("--input", type=str, default=str(DEFAULT_INPUT), help="Ruta a matriz_screening_titulo_resumen.xlsx")
    parser.add_argument("--root", type=str, default=str(DEFAULT_ROOT), help="Carpeta raíz del proyecto")
    parser.add_argument("--top", type=int, default=12, help="Número de candidatos preliminares a exportar")
    args = parser.parse_args()

    root = Path(args.root)
    input_path = Path(args.input)
    if not input_path.exists():
        alt = root / "05_title_abstract_screening" / "matriz_screening_titulo_resumen.xlsx"
        if alt.exists():
            input_path = alt

    dirs = ensure_output_dirs(root)
    df = read_screening_matrix(input_path)
    classified = apply_classification_and_scoring(df)
    classified = assign_top12(classified, top_n=args.top)
    outputs = write_outputs(classified, dirs, top_n=args.top)
    log_path = write_log(dirs, outputs, classified, input_path, top_n=args.top)

    n_excluir = int((classified["clasificacion_asistida"] == "excluir_probable").sum())
    n_incluir = int((classified["clasificacion_asistida"] == "incluir_probable").sum())
    n_dudoso = int((classified["clasificacion_asistida"] == "dudoso_revision_humana").sum())
    n_top = int(classified["estatus_candidato_12"].isin(["candidato_preliminar", "candidato_condicional"]).sum())

    print("Proceso de cribado asistido completado.")
    print(f"Registros únicos procesados: {len(classified)}.")
    print(f"Excluir probable por reglas: {n_excluir}.")
    print(f"Incluir probable por reglas: {n_incluir}.")
    print(f"Dudoso/revisión humana: {n_dudoso}.")
    print(f"Top {args.top} candidatos preliminares generados: {n_top}.")
    print("Archivo principal:")
    print(outputs["matriz_cribado_asistido_trazable"])
    print("\nOtros archivos:")
    for name, path in outputs.items():
        if name != "matriz_cribado_asistido_trazable":
            print(f"- {name}: {path}")
    print(f"- log: {log_path}")
    print("\nAdvertencia:")
    print("Este cribado es asistido por reglas y requiere validación humana. No corresponde a inclusión final PRISMA.")


if __name__ == "__main__":
    main()
